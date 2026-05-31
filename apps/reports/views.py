from django.shortcuts import render
from django.http import HttpResponse
from apps.production.models import WorkEntry
from apps.labor.models import Labor, Payment
from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.contrib.auth.decorators import login_required

@login_required
def home(request):
    return render(request, 'reports/home.html')

@login_required
def export_labor_excel(request):
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Labor Earnings Report"
    
    # Header
    headers = ['Worker Name', 'Phone', 'Total Earned', 'Total Paid', 'Balance']
    ws.append(headers)
    
    # Styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        
    labors = Labor.objects.all()
    for labor in labors:
        # If dates are provided, we could calculate balance for that range,
        # but usually Worker Earnings is a total overview. 
        # For now, we'll keep it as a total snapshot.
        ws.append([
            labor.name,
            labor.phone,
            labor.total_earned,
            labor.total_paid,
            labor.pending_balance
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=labor_report.xlsx'
    wb.save(response)
    return response

@login_required
def export_production_excel(request):
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Production"
    
    headers = ['Date', 'Lot #', 'Operation', 'Color', 'Worker', 'Pieces', 'Rate', 'Total']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        
    entries = WorkEntry.objects.select_related('lot', 'operation', 'labor', 'lot_color').order_by('-work_date')
    
    if start_date:
        entries = entries.filter(work_date__gte=start_date)
    if end_date:
        entries = entries.filter(work_date__lte=end_date)
        
    for e in entries:
        ws.append([
            e.work_date.strftime('%d-%m-%Y'),
            e.lot.lot_number,
            e.operation.name,
            e.lot_color.color_name if e.lot_color else "General",
            e.labor.name,
            e.pieces,
            e.rate,
            e.total_amount
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=production_{start_date or "all"}_to_{end_date or "now"}.xlsx'
    wb.save(response)
    return response
